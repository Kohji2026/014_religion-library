"""
宗教ライブラリ データ変換スクリプト
=====================================
C:\\Users\\kohji\\OneDrive\\01_記録\\22_宗教一覧.xlsx を読み込み、
religion_data.json を生成する。

Excelの構造（宗教一覧シート）:
  行1    : 空白
  行2    : 列B=空白、列C以降=宗教名（28宗教）
  行3-43 : 列B=属性名、列C以降=各宗教の値（転置形式）

実行方法:
    uv run --with openpyxl convert_to_json.py
"""

import openpyxl
import json
import re
import shutil
import tempfile
from pathlib import Path

# ========== 設定 ==========
SCRIPT_DIR  = Path(__file__).parent
INPUT_EXCEL = Path(r"C:\Users\kohji\OneDrive\01_記録\22_宗教一覧.xlsx")
OUTPUT_JSON = SCRIPT_DIR / "religion_data.js"
MEDIA_BASE  = Path(r"C:\Users\kohji\OneDrive\01_記録\頻度少\宗教")

# ========== 属性インデックス（row 2=インデックス0 → 属性値がある行はrow3以降）==========
# main_attrsリスト（0-indexed）から属性名とJSONキーの対応
ATTR_KEYS = [
    ("OneNote（基礎）", "onenote_url"),
    ("OneNote（画像）", "onenote_img"),
    ("NotebookLM",      "notebooklm"),
    ("思想",            "philosophy"),
    ("創始者",          "founder"),
    ("教典",            "scripture"),
    ("場所",            "place"),
    ("聖地",            "holy_place"),
    ("起源",            "origin"),
    ("消滅",            "extinction"),
    ("信者数",          "believers"),
    ("崇拝対象",        "worship_object"),
    ("行事",            "events"),
    ("衣",              "costume"),
    ("食",              "food"),
    ("施設",            "facilities"),
]
# キーワードは1-15、動画は①-⑩、OneDriveは最後
KW_COUNT  = 15
VID_COUNT = 10

# ========== 開宗年を数値に変換 ==========
def estimate_year(origin_str):
    if not origin_str:
        return None
    s = str(origin_str).strip()
    # 紀元前XX世紀 → 負の値（例: B.C.13世紀 → -1250）
    m = re.search(r'B\.?C\.?\s*(\d+)世紀', s)
    if m:
        return -(int(m.group(1)) * 100 - 50)
    m = re.search(r'紀元前(\d+)世紀', s)
    if m:
        return -(int(m.group(1)) * 100 - 50)
    # B.C.XXXX年 → 負
    m = re.search(r'B\.?C\.?\s*(\d+)', s)
    if m:
        return -int(m.group(1))
    # 約XXXX年前 → 2026 - XXXX
    m = re.search(r'約(\d+)年前', s)
    if m:
        return 2026 - int(m.group(1))
    # XX世紀 → XX*100-50
    m = re.search(r'(\d+)世紀', s)
    if m:
        return int(m.group(1)) * 100 - 50
    # 年号（4桁）
    m = re.search(r'\b(1[0-9]{3}|20[0-9]{2})\b', s)
    if m:
        return int(m.group(1))
    return None

# ========== ローカルファイルURL生成 ==========
def local_url(name, ext):
    path = MEDIA_BASE / name / f"{name}.{ext}"
    if path.exists():
        return path.as_uri()
    return ""

# ========== 分類を思想フィールドから推定 ==========
def classify(name, philosophy):
    p = str(philosophy or "").strip()
    n = str(name or "").strip()
    # 思想フィールドによる判定
    if "一神教" in p:
        return "一神教"
    if "多神教" in p:
        return "多神教"
    if "無神" in p:
        return "無神論的宗教"
    if "仏教" in p or "禅" in p:
        return "仏教系"
    # 宗教名による補完
    if n in ("仏教", "禅", "ジャイナ教"):
        return "仏教系"
    if n in ("道教", "儒教", "神道", "バラモン教", "ヒンドゥー教"):
        return "多神教"
    if n in ("ユダヤ教", "キリスト教", "イスラム教", "ゾロアスター教",
             "シク教", "マニ教", "モルモン教", "エホバの証人",
             "統一教会", "サイエントロジー"):
        return "一神教"
    if n in ("創価学会", "天理教", "幸福の科学", "PL教団", "国柱会",
             "立正佼成会", "霊友会", "金光教", "黒住教", "オウム真理教"):
        return "日本の新宗教"
    return "その他"

# ========== セル値を安全に文字列変換 ==========
def sv(val):
    if val is None:
        return ""
    return str(val).strip()

def main():
    print(f"読み込み中: {INPUT_EXCEL}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    shutil.copy2(INPUT_EXCEL, tmp_path)
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
    finally:
        tmp_path.unlink(missing_ok=True)

    ws = wb["宗教一覧"]

    # 行2: 宗教名（col3以降）
    relig_names = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            relig_names.append((c, str(v).strip()))

    religions = []
    for col_idx, name in relig_names:
        # セル値を取得（エラー値除去）
        def get_attr(attr_idx, _col=col_idx):
            row_num = attr_idx + 2
            v = sv(ws.cell(row_num, _col).value)
            return "" if v.startswith("#") else v

        # ハイパーリンクURLを取得
        def get_link(attr_idx, _col=col_idx):
            row_num = attr_idx + 2
            cell = ws.cell(row_num, _col)
            if cell.hyperlink and cell.hyperlink.target:
                t = str(cell.hyperlink.target).strip()
                return t if t.startswith("http") else ""
            return ""

        # 基本属性
        philosophy     = get_attr(3)
        founder        = get_attr(4)
        scripture      = get_attr(5)
        place          = get_attr(6)
        holy_place     = get_attr(7)
        origin         = get_attr(8)
        extinction     = get_attr(9)
        believers      = get_attr(10)
        worship_object = get_attr(11)
        events         = get_attr(12)
        costume        = get_attr(13)
        food           = get_attr(14)
        facilities     = get_attr(15)

        # 動画プレイリスト（attr idx=0 = row2: 宗教名セル自体にYouTubeリンクが設定されている）
        video_playlist_url = get_link(0)

        # NotebookLM（attr idx=2: セル値は"☞宗教名"形式、URLはハイパーリンクから取得）
        notebooklm_url = get_link(2)

        # キーワード（idx 16-30 = 15個）
        keywords = []
        for ki in range(16, 16 + KW_COUNT):
            v = get_attr(ki)
            if v:
                keywords.append(v)

        # 動画（idx 31-40 = 10個）: URLはハイパーリンクから取得
        video_start = 16 + KW_COUNT  # = 31
        videos = []
        for vi in range(video_start, video_start + VID_COUNT):
            url = get_link(vi)
            label = get_attr(vi)
            # "☞ 宗教名" 形式のラベルからテキストを整形
            label_clean = label.replace("☞", "").strip() if label else ""
            if url:
                videos.append({"label": label_clean or f"動画{vi - video_start + 1}", "url": url})

        # OneDrive（idx 41）
        onedrive = get_link(video_start + VID_COUNT)

        # 分類判定
        classification = classify(name, philosophy)
        # 開宗年数値変換
        founded_year = estimate_year(origin)

        religion = {
            "id":              len(religions) + 1,
            "name":            name,
            "classification":  classification,
            "philosophy":      philosophy,
            "founder":         founder,
            "scripture":       scripture,
            "place":           place,
            "holy_place":      holy_place,
            "origin":          origin,
            "founded_year":    founded_year,
            "extinction":      extinction,
            "believers":       believers,
            "worship_object":  worship_object,
            "events":          events,
            "costume":         costume,
            "food":            food,
            "facilities":      facilities,
            "keywords":        keywords,
            "videos":          videos,
            "video_url":       video_playlist_url,
            "notebooklm_url":  notebooklm_url,
            "onedrive":        onedrive,
            "png_url":         local_url(name, "png"),
            "pdf_url":         local_url(name, "pdf"),
        }
        religions.append(religion)

    output = {
        "meta": {
            "title":       "宗教総合ライブラリ",
            "description": "浩二さんのパーソナル宗教ライブラリ（28宗教）",
            "version":     "1.0",
            "count":       len(religions),
        },
        "religions": religions,
    }

    # JSファイルとして出力（fetch不要でChromeローカルファイルでも動作）
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        f.write("const RELIGION_DATA = ")
        json.dump(output, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    cls_counts = {}
    for r in religions:
        cls_counts[r["classification"]] = cls_counts.get(r["classification"], 0) + 1

    print(f"完了！")
    print(f"  宗教数        : {len(religions)} 件")
    for cls, cnt in sorted(cls_counts.items()):
        print(f"    {cls}: {cnt}件")
    print(f"  PNGあり          : {sum(1 for r in religions if r['png_url'])} 件")
    print(f"  PDFあり          : {sum(1 for r in religions if r['pdf_url'])} 件")
    print(f"  動画PLあり        : {sum(1 for r in religions if r['video_url'])} 件")
    print(f"  NotebookLMあり   : {sum(1 for r in religions if r['notebooklm_url'])} 件")
    print(f"  動画リンクあり   : {sum(1 for r in religions if r['videos'])} 件")
    print(f"  出力先        : {OUTPUT_JSON}")

if __name__ == "__main__":
    main()
